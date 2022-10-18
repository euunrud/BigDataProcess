import openpyxl

wb=openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
#
row_id = 1
for row in ws:
        if row_id != 1:
                sum = ws.cell(row=row_id, column=3).value * 0.3
                sum += ws.cell(row=row_id, column=4).value *0.35
                sum += ws.cell(row=row_id, column=5).value *0.34
                sum += ws.cell(row=row_id, column=6).value
                ws.cell(row=row_id, column=7).value = sum
        #       print(sum,"sum")
        row_id += 1

row_id = 1
scores=dict()
for row in ws:
        if row_id != 1:
                score = ws.cell(row=row_id, column=7).value
                if score not in scores:
                        scores[score]=1
                else:
                      mport openpyxl

wb=openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
for row in ws:
        if row_id != 1:
                sum = ws.cell(row=row_id, column=3).value * 0.3
                sum += ws.cell(row=row_id, column=4).value *0.35
                sum += ws.cell(row=row_id, column=5).value *0.34
                sum += ws.cell(row=row_id, column=6).value
                ws.cell(row=row_id, column=7).value = sum
        #       print(sum,"sum")
        row_id += 1

row_id = 1
scores=dict()
for row in ws:
        if row_id != 1:
                score = ws.cell(row=row_id, column=7).value
                if score not in scores:
                        scores[score]=1
                else:
                      import openpyxl

wb=openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
for row in ws:
	if row_id != 1:
		sum = ws.cell(row=row_id, column=3).value * 0.3
		sum += ws.cell(row=row_id, column=4).value *0.35
		sum += ws.cell(row=row_id, column=5).value *0.34
		sum += ws.cell(row=row_id, column=6).value
		ws.cell(row=row_id, column=7).value = sum
	#	print(sum,"sum")
	row_id += 1

row_id = 1
scores=dict()
for row in ws:
	if row_id != 1:
		score = ws.cell(row=row_id, column=7).value
		if score not in scores:
			scores[score]=1
		else:
			scores[score] += 1
	#		print(scores)
	row_id += 1
#print(scores)
scores= sorted(scores.items(), reverse=True)
#print("22",scores)
big = len(scores)
num=0
rslt =dict()
num1=0
num2 =0
num3=0
num4=0
num5=0
row = row_id-2		
#print("row",row)
for i in range(len(scores)):
#	print(i,"i")
#	print("sc",scores[i][1])
	num+=scores[i][1]

	if num <= row * 0.15 and num1 == 0:
		if num + scores[i + 1][1] > row * 0.15:
			num1 = 1
			rslt['A+'] = scores[i][0]
	elif num <= row * 0.3 and num2 == 0:
		if num + scores[i + 1][1] > row * 0.3:
			num2 = 1
			rslt['A0'] = scores[i][0]  
	elif num <= row * 0.5 and num3 == 0:
		if num + scores[i + 1][1] > row * 0.5:
			num3 = 1
			rslt['B+'] = scores[i][0] 
	elif num <= row * 0.7 and num4 == 0:
		if num + scores[i + 1][1] > row * 0.7:
			num4 = 1
			rslt['B0'] = scores[i][0]
	elif num <= row * 0.85 and num5 == 0:
		if num + scores[i + 1][1] > row * 0.85:
			num5 = 1
			rslt['C+'] = scores[i][0]
row_id = 1
#print(rslt)

for row in ws:
	if row_id != 1:
		score = ws.cell(row=row_id, column=7).value
#		print(score, "score--")
		if score >= rslt['A+']:
			ws.cell(row=row_id, column=8).value='A+'
		elif score >= rslt['A0']:
			ws.cell(row=row_id, column=8).value='A0'
		elif score >= rslt['B+']:
			ws.cell(row=row_id, column=8).value='B+'
		elif score >= rslt['B0']:
			ws.cell(row=row_id, column=8).value='B0'
		elif score >= rslt['C+']:
			ws.cell(row=row_id, column=8).value='C+'
		else:
			ws.cell(row=row_id, column=8).value='C0'
	row_id += 1

wb.save("student.xlsx")
